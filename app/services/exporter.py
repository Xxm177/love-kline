import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HEADER_FONT = Font(name="Arial", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
CELL_ALIGNMENT = Alignment(vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin", color="333333"),
    right=Side(style="thin", color="333333"),
    top=Side(style="thin", color="333333"),
    bottom=Side(style="thin", color="333333"),
)


def _style_header(ws, headers: list[str]):
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def _style_data(ws, start_row: int, end_row: int, num_cols: int):
    for row in range(start_row, end_row + 1):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = CELL_ALIGNMENT
            cell.border = THIN_BORDER


def _auto_width(ws, min_width: int = 10, max_width: int = 40):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            val = str(cell.value or "")
            max_len = max(max_len, len(val))
        width = max(min_width, min(max_len + 2, max_width))
        ws.column_dimensions[col_letter].width = width


def export_to_excel(
    messages: list[dict],
    index: list[dict],
    kline: list[dict],
) -> io.BytesIO:
    wb = Workbook()

    # ── Sheet 1: Message Analysis ──
    ws1 = wb.active
    ws1.title = "Message Analysis"
    headers1 = ["time", "sender", "message", "score", "dimension", "reason"]
    _style_header(ws1, headers1)

    for row_idx, msg in enumerate(messages, 2):
        ws1.cell(row=row_idx, column=1, value=msg.get("time", ""))
        ws1.cell(row=row_idx, column=2, value=msg.get("sender", ""))
        ws1.cell(row=row_idx, column=3, value=msg.get("message", ""))
        ws1.cell(row=row_idx, column=4, value=msg.get("score", 0))
        ws1.cell(row=row_idx, column=5, value=msg.get("dimension", ""))
        ws1.cell(row=row_idx, column=6, value=msg.get("reason", ""))

    _style_data(ws1, 2, len(messages) + 1, len(headers1))
    _auto_width(ws1)

    # ── Sheet 2: Relationship Index ──
    ws2 = wb.create_sheet("Relationship Index")
    headers2 = ["time", "index"]
    _style_header(ws2, headers2)

    for row_idx, pt in enumerate(index, 2):
        ws2.cell(row=row_idx, column=1, value=pt.get("time", ""))
        ws2.cell(row=row_idx, column=2, value=pt.get("index", 0))

    _style_data(ws2, 2, len(index) + 1, len(headers2))
    _auto_width(ws2)

    # ── Sheet 3: Kline ──
    ws3 = wb.create_sheet("Kline")
    headers3 = ["date", "open", "high", "low", "close"]
    _style_header(ws3, headers3)

    for row_idx, bar in enumerate(kline, 2):
        ws3.cell(row=row_idx, column=1, value=bar.get("date", ""))
        ws3.cell(row=row_idx, column=2, value=bar.get("open", 0))
        ws3.cell(row=row_idx, column=3, value=bar.get("high", 0))
        ws3.cell(row=row_idx, column=4, value=bar.get("low", 0))
        ws3.cell(row=row_idx, column=5, value=bar.get("close", 0))

    _style_data(ws3, 2, len(kline) + 1, len(headers3))
    _auto_width(ws3)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
